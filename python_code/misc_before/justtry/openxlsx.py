# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb=load_workbook('demo.xlsx')
ws=wb['first']
print ws['A1'].value
